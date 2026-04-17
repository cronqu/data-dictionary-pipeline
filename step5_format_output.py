#!/usr/bin/env python3
"""
STEP 5 — Format AI JSON output into a styled xlsx data dictionary  (AUTOMATED)
=============================================================================
PURPOSE : Reads the AI's JSON response (ai_output.json) and writes a
          professional, formatted Excel workbook (data_dictionary.xlsx).

INPUTS  : ai_output.json          (paste the AI's full response into this file)
          synthetic_template.csv  (used to pull the synthetic example values)

OUTPUT  : data_dictionary.xlsx    (open and review in Stage 6)

STYLING : Rows are color-coded by source system. Colors are assigned
          automatically from a palette based on the distinct source_system
          values returned by the AI — no hard-coded system names required.

PRIVACY : This script NEVER reads your real CSV. It only reads the AI's output
          and your synthetic template.
=============================================================================
"""

import csv
import json
import os
import re
import sys
import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    print("ERROR: openpyxl is not installed.")
    print("Install it with:  pip3 install openpyxl")
    sys.exit(1)

SCRIPT_DIR     = os.path.dirname(os.path.abspath(__file__))
AI_OUTPUT_FILE = os.path.join(SCRIPT_DIR, "ai_output.json")
TEMPLATE_FILE  = os.path.join(SCRIPT_DIR, "synthetic_template.csv")
OUTPUT_FILE    = os.path.join(SCRIPT_DIR, "data_dictionary.xlsx")

# ── Validate inputs ───────────────────────────────────────────────────────────
if not os.path.exists(AI_OUTPUT_FILE):
    print("ERROR: ai_output.json not found.")
    print("Please paste the AI's JSON response into ai_output.json and rerun.")
    sys.exit(1)

# ── Parse AI output ───────────────────────────────────────────────────────────
with open(AI_OUTPUT_FILE, "r", encoding="utf-8") as f:
    raw = f.read().strip()

# Strip markdown code fences if present (```json ... ```)
raw = re.sub(r"^```json\s*", "", raw, flags=re.IGNORECASE)
raw = re.sub(r"^```\s*", "", raw)
raw = re.sub(r"\s*```$", "", raw)
raw = raw.strip()

try:
    records = json.loads(raw)
except json.JSONDecodeError as e:
    print(f"ERROR: Could not parse ai_output.json as JSON.")
    print(f"JSON error: {e}")
    print()
    print("Tips:")
    print("  • Make sure you copied the AI's COMPLETE response")
    print("  • The file should contain only the JSON array (starting with '[')")
    print("  • If the AI wrapped it in ```json ... ``` that's fine — we strip those")
    sys.exit(1)

if not isinstance(records, list):
    print("ERROR: ai_output.json should contain a JSON array (starting with '[').")
    sys.exit(1)

n = len(records)
print(f"Parsed {n} records from ai_output.json")

# ── Required fields ───────────────────────────────────────────────────────────
REQUIRED_KEYS = {"field", "data_type", "source_table", "source_system",
                 "code_desc_pair", "description"}

issues = []
for i, rec in enumerate(records):
    missing = REQUIRED_KEYS - set(rec.keys())
    if missing:
        issues.append(f"  Record {i+1} ({rec.get('field','?')}): missing keys {missing}")
if issues:
    print(f"\nWARNING: {len(issues)} records have missing keys:")
    for msg in issues[:10]:
        print(msg)
    if len(issues) > 10:
        print(f"  ... and {len(issues)-10} more")
    print("\nContinuing — missing fields will be shown as '[MISSING]' in the xlsx.\n")

# ── Read synthetic example values ─────────────────────────────────────────────
example_map = {}
if os.path.exists(TEMPLATE_FILE):
    with open(TEMPLATE_FILE, "r", newline="", encoding="utf-8") as f:
        reader = csv.reader(f)
        header_row  = next(reader, [])
        example_row = next(reader, [])
    for col, val in zip(header_row, example_row):
        example_map[col] = val.strip() if val.strip() else "—"
else:
    print("WARNING: synthetic_template.csv not found. Example values will be blank.")

# ── Color palette — assigned dynamically from distinct source_system values ────
# Colors are soft pastels so text remains readable on every row.
_PALETTE = [
    "D9E2F3",  # light blue
    "E2EFDA",  # light green
    "FFF2CC",  # light yellow
    "FCE4D6",  # light peach/orange
    "E2D9F3",  # light lavender
    "DAEEF3",  # light teal
    "EAD1DC",  # light rose
    "D9F0DA",  # mint
    "F2DCDB",  # light salmon
    "D6E4F0",  # sky blue
    "FDE9D9",  # cream peach
    "E8F5E9",  # pale green
]

# Collect distinct source_system values in the order they first appear
_seen_systems: list = []
for _rec in records:
    _sys = (_rec.get("source_system") or "—").strip()
    if _sys not in _seen_systems:
        _seen_systems.append(_sys)

section_fills = {
    sys_name: PatternFill(
        start_color=_PALETTE[i % len(_PALETTE)],
        end_color=_PALETTE[i % len(_PALETTE)],
        fill_type="solid",
    )
    for i, sys_name in enumerate(_seen_systems)
}

# ── Build workbook ────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Data Descriptions"

# Styling constants
header_font  = Font(bold=True, size=11, color="FFFFFF")
header_fill  = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
data_font    = Font(size=10)
thin_border  = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)

HEADER_LABELS = [
    "Field",
    "Example Value (Synthetic)",
    "Source Table",
    "Data Type",
    "Source System",
    "Code/Desc Pair",
    "Description & Comments",
]

# Write header row
for col_idx, label in enumerate(HEADER_LABELS, 1):
    cell = ws.cell(row=1, column=col_idx, value=label)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border

# Write data rows
for row_idx, rec in enumerate(records, 2):
    field       = rec.get("field", "[MISSING]")
    data_type   = rec.get("data_type", "[MISSING]")
    source_tbl  = rec.get("source_table", "—")
    source_sys  = rec.get("source_system", "—")
    pair        = rec.get("code_desc_pair", "—")
    description = rec.get("description", "[MISSING]")
    example_val = example_map.get(field, "—")

    values = [field, example_val, source_tbl, data_type, source_sys, pair, description]

    # Determine row fill from source_system
    row_fill = section_fills.get((source_sys or "—").strip())

    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.font = data_font
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = thin_border
        # Apply section color to first 2 cols (Field + Example)
        if row_fill and col_idx <= 2:
            cell.fill = row_fill

# Column widths
ws.column_dimensions["A"].width = 32
ws.column_dimensions["B"].width = 38
ws.column_dimensions["C"].width = 18
ws.column_dimensions["D"].width = 22
ws.column_dimensions["E"].width = 14
ws.column_dimensions["F"].width = 26
ws.column_dimensions["G"].width = 85

ws.freeze_panes = "A2"
ws.sheet_view.showGridLines = True

# ── Summary sheet ─────────────────────────────────────────────────────────────
ws2 = wb.create_sheet("Summary")
bold = Font(bold=True)
now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")

# Build a dynamic legend: for each source_system, list its distinct source_tables
_system_tables: dict = {}
for _rec in records:
    _sys = (_rec.get("source_system") or "—").strip()
    _tbl = (_rec.get("source_table") or "—").strip()
    _system_tables.setdefault(_sys, set()).add(_tbl)

summary_rows = [
    ("File described:", "data file (see ai_input.txt)"),
    ("Data dictionary generated:", now_str),
    ("Total columns described:", n),
    ("", ""),
    ("PRIVACY NOTE:", "Synthetic example values only — no real participant data included."),
    ("", ""),
    ("Source System Legend:", "Source Tables"),
]

for _sys in _seen_systems:
    _tables = sorted(
        t for t in _system_tables.get(_sys, set()) if t not in ("—", "")
    )
    _table_str = ", ".join(_tables) if _tables else "—"
    summary_rows.append((_sys, _table_str))

for row_idx, (key, val) in enumerate(summary_rows, 1):
    c1 = ws2.cell(row=row_idx, column=1, value=key)
    c2 = ws2.cell(row=row_idx, column=2, value=val)
    if key and not str(val).startswith("(") and row_idx <= 7:
        c1.font = bold
    if key in section_fills:
        c1.fill = section_fills[key]

ws2.column_dimensions["A"].width = 22
ws2.column_dimensions["B"].width = 85

# ── Save ──────────────────────────────────────────────────────────────────────
wb.save(OUTPUT_FILE)
print(f"✅  Saved: {OUTPUT_FILE}")
print(f"   {n} rows | 7 columns | color-coded by source table")
print()
print("──────────────────────────────────────────────────────────────────")
print("NEXT STEP (Stage 6 — Manual):")
print("──────────────────────────────────────────────────────────────────")
print()
print("1. Open data_dictionary.xlsx")
print("2. Review all descriptions — correct any errors or gaps the AI made")
print("3. Check the 'Code/Desc Pair' column for any missed linkages")
print("4. Add any custom notes in the 'Description & Comments' column")
print("5. Save the final file")
print()
print("Your data dictionary is complete! ✅")
print()
